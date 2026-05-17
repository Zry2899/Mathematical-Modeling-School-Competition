"""
Build histogram bin counts and Excel column charts for each raw loss component.

Outputs:
    result/task2_raw_loss_histogram_bins.csv
    result/task2_raw_loss_histograms.xlsx
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


ROOT = Path(__file__).resolve().parent.parent
RESULT_DIR = ROOT / "result"


def main() -> None:
    event = pd.read_csv(RESULT_DIR / "task2_raw_loss_history.csv", encoding="utf-8-sig")
    monthly = pd.read_csv(RESULT_DIR / "task2_raw_loss_monthly_history.csv", encoding="utf-8-sig")

    series_map = {
        "abs_LC_raw": event["LC_raw"].abs(),
        "LF_raw": event["LF_raw"],
        "LV_raw": event["LV_raw"],
        "LT_raw": event["LT_raw"],
        "LE_raw": event["LE_raw"],
        "Lpi_raw": monthly["Lpi_raw"],
    }

    all_bins = []
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name, series in series_map.items():
        hist = build_histogram(name, series, bins=12)
        all_bins.append(hist)
        add_histogram_sheet(workbook, name, hist)

    out = pd.concat(all_bins, ignore_index=True)
    out.to_csv(
        RESULT_DIR / "task2_raw_loss_histogram_bins.csv",
        index=False,
        encoding="utf-8-sig",
    )
    workbook.save(RESULT_DIR / "task2_raw_loss_histograms.xlsx")

    print("Generated histogram outputs:")
    print("- task2_raw_loss_histogram_bins.csv")
    print("- task2_raw_loss_histograms.xlsx")
    print()
    for name in series_map:
        print(f"\n{name}")
        print(out[out["loss"] == name][["bin_label", "count"]].to_string(index=False))


def build_histogram(name: str, series: pd.Series, bins: int = 12) -> pd.DataFrame:
    values = pd.to_numeric(series, errors="coerce").dropna().to_numpy(dtype=float)
    values = values[np.isfinite(values)]
    if len(values) == 0:
        return pd.DataFrame(columns=["loss", "bin_left", "bin_right", "bin_label", "count"])

    min_value = float(np.min(values))
    max_value = float(np.max(values))
    if np.isclose(min_value, max_value):
        edges = np.array([min_value, max_value + 1.0])
    else:
        edges = np.linspace(min_value, max_value, bins + 1)

    counts, edges = np.histogram(values, bins=edges)
    rows = []
    for idx, count in enumerate(counts):
        left = float(edges[idx])
        right = float(edges[idx + 1])
        label = f"[{left:.6g}, {right:.6g}{']' if idx == len(counts)-1 else ')'}"
        rows.append(
            {
                "loss": name,
                "bin_left": left,
                "bin_right": right,
                "bin_label": label,
                "count": int(count),
            }
        )
    return pd.DataFrame(rows)


def add_histogram_sheet(workbook: Workbook, name: str, hist: pd.DataFrame) -> None:
    safe_name = name[:31]
    ws = workbook.create_sheet(safe_name)
    for row in dataframe_to_rows(hist[["bin_label", "count"]], index=False, header=True):
        ws.append(row)

    chart = BarChart()
    chart.title = f"{name} frequency by interval"
    chart.x_axis.title = "Interval"
    chart.y_axis.title = "Count"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "D2")


if __name__ == "__main__":
    main()
