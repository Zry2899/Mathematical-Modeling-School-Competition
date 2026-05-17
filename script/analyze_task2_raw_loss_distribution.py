"""
Export raw task-2 loss components under the historical adjustment path and
plot their distributions before choosing a normalization scheme.

Outputs:
    result/task2_raw_loss_history.csv
    result/task2_raw_loss_monthly_history.csv
    result/task2_raw_loss_distribution_summary.csv
    result/task2_raw_loss_distribution.xlsx
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.utils.dataframe import dataframe_to_rows

from task2_social_welfare_loss import get_weight_scenario, load_parameters, social_welfare_loss


ROOT = Path(__file__).resolve().parent.parent
RESULT_DIR = ROOT / "result"
LOSS_COLUMNS = ["LC_raw", "LF_raw", "LV_raw", "LT_raw", "LE_raw"]


def main() -> None:
    data = pd.read_csv(RESULT_DIR / "task2_event_model_input_clean.csv", encoding="utf-8-sig")
    params = load_parameters()
    weights = get_weight_scenario("equal_weight")

    alpha_g = np.where(
        data["f_gasoline"].abs() > 1e-12,
        data["u_hist_gasoline"] / data["f_gasoline"],
        0.0,
    )
    alpha_d = np.where(
        data["f_diesel"].abs() > 1e-12,
        data["u_hist_diesel"] / data["f_diesel"],
        0.0,
    )
    alpha_g = np.clip(alpha_g, 0.0, 1.0)
    alpha_d = np.clip(alpha_d, 0.0, 1.0)

    result = social_welfare_loss(
        event_data=data,
        parameters=params,
        preference_weights=weights,
        alpha_gasoline=alpha_g,
        alpha_diesel=alpha_d,
    )

    event_loss = result["event_loss"].copy()
    monthly_loss = result["monthly_loss"].copy()

    # Reattach historical adjustment columns for easier inspection.
    event_loss["u_hist_gasoline"] = data["u_hist_gasoline"].values
    event_loss["u_hist_diesel"] = data["u_hist_diesel"].values
    event_loss["f_gasoline"] = data["f_gasoline"].values
    event_loss["f_diesel"] = data["f_diesel"].values

    lpi_by_month = monthly_loss[["month", "Lpi_raw"]].copy()
    lpi_by_month["month"] = pd.to_datetime(lpi_by_month["month"])

    event_loss.to_csv(
        RESULT_DIR / "task2_raw_loss_history.csv",
        index=False,
        encoding="utf-8-sig",
    )
    monthly_loss.to_csv(
        RESULT_DIR / "task2_raw_loss_monthly_history.csv",
        index=False,
        encoding="utf-8-sig",
    )

    summary = build_summary(event_loss, monthly_loss)
    summary.to_csv(
        RESULT_DIR / "task2_raw_loss_distribution_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    write_distribution_workbook(event_loss, monthly_loss, summary)

    print("Generated raw loss distribution outputs:")
    print("- task2_raw_loss_history.csv")
    print("- task2_raw_loss_monthly_history.csv")
    print("- task2_raw_loss_distribution_summary.csv")
    print("- task2_raw_loss_distribution.xlsx")
    print()
    print(summary.to_string(index=False))


def build_summary(event_loss: pd.DataFrame, monthly_loss: pd.DataFrame) -> pd.DataFrame:
    series_map = {
        "LC_raw": event_loss["LC_raw"],
        "abs_LC_raw": event_loss["LC_raw"].abs(),
        "LF_raw": event_loss["LF_raw"],
        "LV_raw": event_loss["LV_raw"],
        "LT_raw": event_loss["LT_raw"],
        "LE_raw": event_loss["LE_raw"],
        "Lpi_raw": monthly_loss["Lpi_raw"],
    }

    rows = []
    for name, series in series_map.items():
        values = pd.to_numeric(series, errors="coerce").dropna()
        rows.append(
            {
                "loss": name,
                "n": len(values),
                "mean": values.mean(),
                "mean_abs": values.abs().mean(),
                "std": values.std(ddof=1),
                "min": values.min(),
                "p25": values.quantile(0.25),
                "median": values.quantile(0.50),
                "p75": values.quantile(0.75),
                "p90": values.quantile(0.90),
                "p95": values.quantile(0.95),
                "max": values.max(),
                "zero_ratio": float((values.abs() < 1e-12).mean()),
            }
        )
    return pd.DataFrame(rows)


def write_distribution_workbook(
    event_loss: pd.DataFrame,
    monthly_loss: pd.DataFrame,
    summary: pd.DataFrame,
) -> None:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "summary"

    for row in dataframe_to_rows(summary, index=False, header=True):
        ws_summary.append(row)

    chart = BarChart()
    chart.title = "Raw Loss P90 by Component"
    chart.y_axis.title = "P90 raw loss"
    chart.x_axis.title = "Loss component"
    data = Reference(ws_summary, min_col=10, min_row=1, max_row=ws_summary.max_row)
    cats = Reference(ws_summary, min_col=1, min_row=2, max_row=ws_summary.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws_summary.add_chart(chart, "O2")

    loss_wide = pd.DataFrame(
        {
            "date": event_loss["date"],
            "abs_LC_raw": event_loss["LC_raw"].abs(),
            "LF_raw": event_loss["LF_raw"],
            "LV_raw": event_loss["LV_raw"],
            "LT_raw": event_loss["LT_raw"],
            "LE_raw": event_loss["LE_raw"],
        }
    )
    ws_event = workbook.create_sheet("event_raw_loss")
    for row in dataframe_to_rows(loss_wide, index=False, header=True):
        ws_event.append(row)

    ws_month = workbook.create_sheet("monthly_lpi")
    monthly_keep = monthly_loss[
        [
            "month",
            "x_gasoline",
            "x_diesel",
            "cpi_yoy",
            "cpi_hat_next",
            "inflation_excess",
            "deflation_gap",
            "Lpi_raw",
        ]
    ].copy()
    for row in dataframe_to_rows(monthly_keep, index=False, header=True):
        ws_month.append(row)

    ws_sorted = workbook.create_sheet("sorted_distributions")
    sorted_df = build_sorted_distribution_table(event_loss, monthly_loss)
    for row in dataframe_to_rows(sorted_df, index=False, header=True):
        ws_sorted.append(row)

    chart2 = ScatterChart()
    chart2.title = "Sorted Raw Loss Distributions"
    chart2.x_axis.title = "Sorted index"
    chart2.y_axis.title = "Raw loss"
    chart2.height = 9
    chart2.width = 20
    for col_idx in range(2, ws_sorted.max_column + 1):
        xvalues = Reference(ws_sorted, min_col=1, min_row=2, max_row=ws_sorted.max_row)
        yvalues = Reference(ws_sorted, min_col=col_idx, min_row=2, max_row=ws_sorted.max_row)
        series = Series(yvalues, xvalues, title=ws_sorted.cell(row=1, column=col_idx).value)
        chart2.series.append(series)
    ws_sorted.add_chart(chart2, "I2")

    workbook.save(RESULT_DIR / "task2_raw_loss_distribution.xlsx")


def build_sorted_distribution_table(
    event_loss: pd.DataFrame,
    monthly_loss: pd.DataFrame,
) -> pd.DataFrame:
    series_map = {
        "abs_LC_raw": event_loss["LC_raw"].abs(),
        "LF_raw": event_loss["LF_raw"],
        "LV_raw": event_loss["LV_raw"],
        "LT_raw": event_loss["LT_raw"],
        "LE_raw": event_loss["LE_raw"],
        "Lpi_raw": monthly_loss["Lpi_raw"],
    }
    max_len = max(len(s.dropna()) for s in series_map.values())
    out = pd.DataFrame({"rank": np.arange(1, max_len + 1)})
    for name, series in series_map.items():
        values = np.sort(pd.to_numeric(series, errors="coerce").dropna().to_numpy())
        padded = np.full(max_len, np.nan)
        padded[: len(values)] = values
        out[name] = padded
    return out


if __name__ == "__main__":
    main()
